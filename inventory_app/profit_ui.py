from PySide6 import QtWidgets, QtCore, QtGui
import json
import os
from collections import defaultdict
import logging

from openpyxl import Workbook, load_workbook

from matplotlib.figure import Figure
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas

from utils import get_data_dir, load_json_file
from ui_theme import make_button, make_card, frame, label, treeview, COLOR_SUCCESS, COLOR_DANGER, COLOR_PRIMARY, COLOR_TEXT_MUTED, styled_label

SALES_FILE = os.path.join(get_data_dir(), "sales.json")


def load_sales():
    return load_json_file(SALES_FILE, default=[])

def calculate_monthly_data():
    sales = load_sales()
    monthly = defaultdict(lambda: {"revenue": 0, "cost": 0, "profit": 0})
    for sale in sales:
        month = sale.get("month", "Unknown")
        revenue = sale.get("selling_price", 0) * sale.get("quantity", 0)
        cost = sale.get("purchase_price", 0) * sale.get("quantity", 0)
        monthly[month]["revenue"] += revenue
        monthly[month]["cost"] += cost
        monthly[month]["profit"] += (revenue - cost)
    return monthly

def export_to_excel(monthly_data, parent_widget=None):
    if not monthly_data:
        QtWidgets.QMessageBox.critical(parent_widget, "Error", "No data to export")
        return
    try:
        template_name = "monthly_profit_report.xlsx"
        template_path = template_name
        if not os.path.exists(template_path):
            template_path = os.path.join(os.path.dirname(__file__), template_name)

        if os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
            if ws.max_row > 1: ws.delete_rows(2, ws.max_row)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Monthly Profit Report"
            ws.append(["Month", "Revenue", "Cost", "Profit"])

        for month, data in sorted(monthly_data.items()):
            ws.append([month, round(data["revenue"], 2), round(data["cost"], 2), round(data["profit"], 2)])

        desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
        file_name = "monthly_profit_report_exported.xlsx"
        save_path = os.path.join(desktop, file_name)

        wb.save(save_path)
        logging.info(f"Profit report exported to Desktop: {save_path}")
        QtWidgets.QMessageBox.information(parent_widget, "Success", f"Report exported to Desktop:\n {file_name}")
    except Exception:
        logging.error("Failed to export profit report", exc_info=True)
        QtWidgets.QMessageBox.critical(parent_widget, "Error", "Failed to export report")

def create_profit_tab(parent):
    """
    Creates the profit analysis tab, including a table of monthly data and a visual chart.
    """
    monthly_data = calculate_monthly_data()

    # --- Summary Data ---
    total_rev = sum(d["revenue"] for d in monthly_data.values())
    total_cost = sum(d["cost"] for d in monthly_data.values())
    total_prof = sum(d["profit"] for d in monthly_data.values())

    window = QtWidgets.QWidget()
    main_layout = QtWidgets.QVBoxLayout(window)
    main_layout.setContentsMargins(20, 20, 20, 20)
    main_layout.setSpacing(10)

    heading_label = label(window, "PROFITABILITY ANALYSIS", kind="heading")
    heading_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
    main_layout.addWidget(heading_label)

    # --- Summary Cards Frame ---
    summary_frame = QtWidgets.QWidget()
    summary_layout = QtWidgets.QHBoxLayout(summary_frame)
    summary_layout.setContentsMargins(0, 10, 0, 20)
    summary_layout.setSpacing(15)
    main_layout.addWidget(summary_frame)

    summary_labels = {}

    def stats_card(parent, title, value, color):
        card = make_card(parent, padding=20)
        card_layout = QtWidgets.QVBoxLayout(card)
        card_layout.setContentsMargins(10, 10, 10, 10)
        title_lbl = styled_label(card, title, font=("Segoe UI", 14, "bold"))
        title_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        card_layout.addWidget(title_lbl)
        val_lbl = styled_label(card, f"Rs. {value:,.2f}", font=("Segoe UI", 24, "bold"), foreground=color)
        val_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        card_layout.addWidget(val_lbl)
        card_layout.addStretch()
        return val_lbl

    rev_lbl = stats_card(summary_frame, "TOTAL REVENUE", 0, COLOR_PRIMARY)
    cost_lbl = stats_card(summary_frame, "TOTAL COST OF GOODS", 0, COLOR_TEXT_MUTED)
    prof_lbl = stats_card(summary_frame, "NET PROFIT", 0, COLOR_SUCCESS)
    summary_labels = {"rev": rev_lbl, "cost": cost_lbl, "prof": prof_lbl}

    # Controls: month selector / view mode
    ctrl_frame = QtWidgets.QWidget()
    ctrl_layout = QtWidgets.QHBoxLayout(ctrl_frame)
    ctrl_layout.setContentsMargins(0, 0, 0, 10)
    main_layout.addWidget(ctrl_frame)

    months_all = sorted(list({s.get('month') for s in load_sales() if s.get('month')}), reverse=True)
    month_cb = QtWidgets.QComboBox()
    month_cb.addItems(months_all)
    month_cb.setFixedWidth(160)
    ctrl_layout.addWidget(month_cb)

    def set_month_list():
        vals = sorted(list({s.get('month') for s in load_sales() if s.get('month')}), reverse=True)
        month_cb.clear()
        month_cb.addItems(vals)

    def show_month():
        m = month_cb.currentText().strip()
        if not m:
            QtWidgets.QMessageBox.information(window, 'Select month', 'Please select a month (YYYY-MM)')
            return
        populate_month(m)

    show_btn = make_button(ctrl_frame, 'Show Month', command=show_month, kind='primary')
    ctrl_layout.addWidget(show_btn)
    summary_btn = make_button(ctrl_frame, 'Summary', command=lambda: populate_summary(), kind='secondary')
    ctrl_layout.addWidget(summary_btn)
    ctrl_layout.addStretch()

    # Chart area
    chart_objs = {}
    chart_frame = make_card(window, padding=20)
    chart_layout = QtWidgets.QVBoxLayout(chart_frame)
    main_layout.addWidget(chart_frame)
    fig = Figure(figsize=(10, 4), dpi=100)
    ax = fig.add_subplot(111)
    canvas = FigureCanvas(fig)
    chart_layout.addWidget(canvas)
    chart_objs = {'fig': fig, 'ax': ax, 'canvas': canvas}

    # Table
    table_heading = label(window, "MONTHLY PERFORMANCE DATA", kind="heading")
    table_heading.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft)
    main_layout.addWidget(table_heading)

    table_frame = make_card(window, padx=10, pady=10)
    table_layout = QtWidgets.QHBoxLayout(table_frame)
    table_layout.setContentsMargins(0, 0, 0, 0)
    main_layout.addWidget(table_frame)

    columns = ("month", "revenue", "cost", "profit")
    tree = QtWidgets.QTableWidget()
    tree.setColumnCount(len(columns))
    tree.setHorizontalHeaderLabels([col.upper() for col in columns])
    tree.horizontalHeader().setStretchLastSection(True)
    tree.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.Interactive)
    tree.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)

    column_map = {
        "month": ("Month", 250),
        "revenue": ("Revenue", 200),
        "cost": ("Cost", 200),
        "profit": ("Profit", 200)
    }

    for i, (col, (label_text, width)) in enumerate(column_map.items()):
        tree.setHorizontalHeaderItem(i, QtWidgets.QTableWidgetItem(label_text.upper()))
        tree.setColumnWidth(i, width)

    tree.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
    table_layout.addWidget(tree)

    def populate_summary():
        md = calculate_monthly_data()
        # update summary cards
        total_rev = sum(d["revenue"] for d in md.values())
        total_cost = sum(d["cost"] for d in md.values())
        total_prof = sum(d["profit"] for d in md.values())
        rev_lbl.setText(f"Rs. {total_rev:,.2f}")
        cost_lbl.setText(f"Rs. {total_cost:,.2f}")
        prof_lbl.setText(f"Rs. {total_prof:,.2f}")

        # populate table
        tree.setRowCount(0)
        set_month_list()
        for row_idx, (month, data) in enumerate(sorted(md.items(), reverse=True)):
            tree.insertRow(row_idx)
            tree.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(month))
            tree.setItem(row_idx, 1, QtWidgets.QTableWidgetItem(f"{data['revenue']:,.2f}"))
            tree.setItem(row_idx, 2, QtWidgets.QTableWidgetItem(f"{data['cost']:,.2f}"))
            tree.setItem(row_idx, 3, QtWidgets.QTableWidgetItem(f"{data['profit']:,.2f}"))

        # update chart
        try:
            if chart_objs:
                ax = chart_objs['ax']
                canvas = chart_objs['canvas']
                ax.clear()
                months = sorted(md.keys())
                profits = [md[m]['profit'] for m in months]
                colors = [COLOR_SUCCESS if p >= 0 else COLOR_DANGER for p in profits]
                x = list(range(len(months)))
                ax.bar(x, profits, color=colors, edgecolor='black', alpha=0.9)
                ax.set_xticks(x)
                ax.set_xticklabels(months, rotation=45, ha='right')
                ax.set_title('Monthly Net Profit', fontsize=16)
                ax.set_ylabel('Profit (Rs.)', fontsize=12)
                ax.grid(axis='y', linestyle='--', alpha=0.4)
                chart_objs['fig'].tight_layout()
                canvas.draw()
        except Exception:
            logging.exception('Failed to draw summary chart')

    def populate_month(month):
        sales = load_sales()
        rows = [s for s in sales if s.get('month') == month]
        revenue = sum(s.get('selling_price', 0) * s.get('quantity', 0) for s in rows)
        cost = sum(s.get('purchase_price', 0) * s.get('quantity', 0) for s in rows)
        profit = revenue - cost

        # update summary cards for this month
        rev_lbl.setText(f"Rs. {revenue:,.2f}")
        cost_lbl.setText(f"Rs. {cost:,.2f}")
        prof_lbl.setText(f"Rs. {profit:,.2f}")

        # reconfigure table for detail view
        detail_map = {
            "date": ("Date", 140),
            "model": ("Model", 300),
            "qty": ("Qty", 80),
            "selling_price": ("Unit Price", 120),
            "purchase_price": ("Cost", 120),
            "profit": ("Profit", 120)
        }

        tree.setColumnCount(len(detail_map))
        tree.setHorizontalHeaderLabels([v[0].upper() for v in detail_map.values()])
        for i, (col, (label_text, width)) in enumerate(detail_map.items()):
            tree.setHorizontalHeaderItem(i, QtWidgets.QTableWidgetItem(label_text.upper()))
            tree.setColumnWidth(i, width)

        tree.setRowCount(0)
        for row_idx, s in enumerate(rows):
            p = (s.get('selling_price', 0) - s.get('purchase_price', 0)) * s.get('quantity', 0)
            tree.insertRow(row_idx)
            tree.setItem(row_idx, 0, QtWidgets.QTableWidgetItem(str(s.get('date', ''))))
            tree.setItem(row_idx, 1, QtWidgets.QTableWidgetItem(str(s.get('model', ''))))
            tree.setItem(row_idx, 2, QtWidgets.QTableWidgetItem(str(s.get('quantity', ''))))
            tree.setItem(row_idx, 3, QtWidgets.QTableWidgetItem(f"{s.get('selling_price',0):,.2f}"))
            tree.setItem(row_idx, 4, QtWidgets.QTableWidgetItem(f"{s.get('purchase_price',0):,.2f}"))
            tree.setItem(row_idx, 5, QtWidgets.QTableWidgetItem(f"{p:,.2f}"))

        # update chart to show per-sale profit
        try:
            if chart_objs:
                ax = chart_objs['ax']
                canvas = chart_objs['canvas']
                ax.clear()
                labels = [f"{r.get('date')}\n{r.get('model')}" for r in rows]
                profits = [ (r.get('selling_price',0)-r.get('purchase_price',0))*r.get('quantity',0) for r in rows]
                colors = [COLOR_SUCCESS if p >= 0 else COLOR_DANGER for p in profits]
                x = list(range(len(labels)))
                ax.bar(x, profits, color=colors, edgecolor='black', alpha=0.9)
                ax.set_xticks(x)
                ax.set_xticklabels(labels, rotation=45, ha='right')
                ax.set_title(f'Profits for {month}', fontsize=14)
                ax.grid(axis='y', linestyle='--', alpha=0.4)
                chart_objs['fig'].tight_layout()
                canvas.draw()
        except Exception:
            logging.exception('Failed to draw month chart')

    # initial populate
    populate_summary()

    def refresh_data(event=None):
        try:
            populate_summary()
        except Exception:
            logging.exception("Failed to refresh profit data")

    # FOOTER
    footer = QtWidgets.QWidget()
    footer_layout = QtWidgets.QHBoxLayout(footer)
    footer_layout.addStretch()
    main_layout.addWidget(footer)

    export_btn = make_button(footer, "Export to Excel", command=lambda: export_to_excel(calculate_monthly_data(), window), kind="primary")
    footer_layout.addWidget(export_btn)

    return window
