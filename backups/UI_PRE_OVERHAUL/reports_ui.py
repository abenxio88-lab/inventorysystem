"""
Advanced Reports Module
========================
Professional reporting system with PDF, Excel, and PowerPoint exports.
Custom report builder and analytics dashboard.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
from datetime import datetime, timedelta
from typing import Optional, Dict, List

try:
    from .database import get_db_cursor, get_connection
    from .ui_theme import make_card, styled_label, make_button, FONT_REGULAR, FONT_BOLD, COLOR_PRIMARY, COLOR_SUCCESS, COLOR_DANGER, COLOR_WARNING
except (ImportError, ModuleNotFoundError):
    from database import get_db_cursor, get_connection
    from ui_theme import make_card, styled_label, make_button, FONT_REGULAR, FONT_BOLD, COLOR_PRIMARY, COLOR_SUCCESS, COLOR_DANGER, COLOR_WARNING

# Try to import export libraries
try:
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


def create_reports_tab(parent, current_user=None):
    """
    Creates the advanced reports tab.
    """
    window = ttk.Frame(parent, padding=15)
    
    # Header
    header_frame = ttk.Frame(window)
    header_frame.pack(fill="x", pady=(0, 15))
    
    styled_label(header_frame, "📊 Advanced Reports", font=FONT_BOLD).pack(side=tk.LEFT)
    
    # Report type selector
    report_var = tk.StringVar(value="stock_summary")
    
    report_types = [
        ("stock_summary", "📦 Stock Summary"),
        ("low_stock", "⚠️ Low Stock Report"),
        ("sales_daily", "💰 Daily Sales"),
        ("sales_monthly", "📈 Monthly Sales"),
        ("profit_analysis", "💵 Profit Analysis"),
        ("supplier_performance", "🏭 Supplier Performance"),
        ("inventory_valuation", "💼 Inventory Valuation"),
        ("slow_movers", "🐌 Slow Moving Items"),
        ("fast_movers", "🚀 Fast Moving Items"),
    ]
    
    type_frame = ttk.LabelFrame(header_frame, text="Report Type", padding=10)
    type_frame.pack(side=tk.RIGHT)
    
    report_combo = ttk.Combobox(type_frame, textvariable=report_var, values=[r[1] for r in report_types], state="readonly", width=25)
    report_combo.pack()
    report_combo.set("📦 Stock Summary")
    
    # Map display names to IDs
    window.report_map = {name: id for id, name in report_types}
    
    # Date range
    date_frame = ttk.LabelFrame(window, text="Date Range", padding=10)
    date_frame.pack(fill="x", pady=(0, 10))
    
    styled_label(date_frame, "From:").grid(row=0, column=0, padx=5, pady=5)
    from_var = tk.StringVar(value=(datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    from_entry = ttk.Entry(date_frame, textvariable=from_var, width=12)
    from_entry.grid(row=0, column=1, padx=5, pady=5)
    
    styled_label(date_frame, "To:").grid(row=0, column=2, padx=5, pady=5)
    to_var = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
    to_entry = ttk.Entry(date_frame, textvariable=to_var, width=12)
    to_entry.grid(row=0, column=3, padx=5, pady=5)
    
    # Generate button
    def generate_report():
        report_type = window.report_map.get(report_var.get())
        date_from = from_var.get()
        date_to = to_var.get()
        
        generate_report_data(report_type, date_from, date_to, results_frame, current_user)
    
    make_button(date_frame, "🔄 Generate Report", command=generate_report, kind="primary").grid(row=0, column=4, padx=10)
    
    # Export buttons
    export_frame = ttk.Frame(window)
    export_frame.pack(fill="x", pady=(0, 10))
    
    def export_excel():
        report_type = window.report_map.get(report_var.get())
        export_report_to_excel(report_type, results_frame)
    
    def export_pdf():
        report_type = window.report_map.get(report_var.get())
        export_report_to_pdf(report_type, results_frame)
    
    make_button(export_frame, "📄 Export to Excel", command=export_excel, kind="success").pack(side=tk.LEFT, padx=5)
    make_button(export_frame, "📑 Export to PDF", command=export_pdf, kind="danger").pack(side=tk.LEFT, padx=5)
    
    # Results frame
    results_frame = ttk.LabelFrame(window, text="Report Results", padding=10)
    results_frame.pack(fill="both", expand=True)
    
    # Initial message
    styled_label(results_frame, "Select a report type and click Generate", font=FONT_REGULAR).pack(pady=50)
    
    return window


def generate_report_data(report_type: str, date_from: str, date_to: str, parent, current_user=None):
    """Generate report data and display in table."""
    # Clear existing content
    for widget in parent.winfo_children():
        widget.destroy()
    
    try:
        with get_db_cursor() as cur:
            if report_type == 'stock_summary':
                # Stock Summary Report
                cur.execute("""
                    SELECT p.model, p.category, p.sku, p.stock, 
                           p.purchase_price, p.selling_price,
                           (p.stock * p.purchase_price) as stock_value,
                           (p.stock * p.selling_price) as retail_value
                    FROM products p
                    WHERE p.status = 'active'
                    ORDER BY p.model
                """)
                columns = ["Model", "Category", "SKU", "Stock", "Cost Price", "Selling Price", "Stock Value", "Retail Value"]
                title = "Stock Summary Report"
                
            elif report_type == 'low_stock':
                # Low Stock Report
                cur.execute("""
                    SELECT p.model, p.category, p.stock, p.reorder_point, p.min_stock,
                           s.name as supplier_name, s.phone as supplier_phone
                    FROM products p
                    LEFT JOIN suppliers s ON p.supplier_id = s.id
                    WHERE p.stock <= p.reorder_point
                    AND p.status = 'active'
                    ORDER BY p.stock ASC
                """)
                columns = ["Model", "Category", "Current Stock", "Reorder Point", "Min Stock", "Supplier", "Supplier Phone"]
                title = "Low Stock Alert Report"
                
            elif report_type == 'sales_daily':
                # Daily Sales Report
                cur.execute("""
                    SELECT date, model, quantity, selling_price, total
                    FROM sales
                    WHERE date BETWEEN ? AND ?
                    ORDER BY date DESC
                """, (date_from, date_to))
                columns = ["Date", "Model", "Quantity", "Unit Price", "Total"]
                title = f"Daily Sales Report ({date_from} to {date_to})"
                
            elif report_type == 'sales_monthly':
                # Monthly Sales Report
                cur.execute("""
                    SELECT month, 
                           COUNT(*) as transaction_count,
                           SUM(quantity) as total_units,
                           SUM(total) as total_revenue
                    FROM sales
                    WHERE month BETWEEN ? AND ?
                    GROUP BY month
                    ORDER BY month DESC
                """, (date_from, date_to))
                columns = ["Month", "Transactions", "Units Sold", "Total Revenue"]
                title = f"Monthly Sales Report ({date_from} to {date_to})"
                
            elif report_type == 'profit_analysis':
                # Profit Analysis
                cur.execute("""
                    SELECT p.model, p.category,
                           SUM(s.quantity) as units_sold,
                           SUM(s.total) as revenue,
                           SUM(s.quantity * p.purchase_price) as cost,
                           SUM(s.total - (s.quantity * p.purchase_price)) as profit
                    FROM sales s
                    JOIN products p ON s.model = p.model
                    WHERE s.date BETWEEN ? AND ?
                    GROUP BY p.model, p.category
                    ORDER BY profit DESC
                """, (date_from, date_to))
                columns = ["Model", "Category", "Units Sold", "Revenue", "Cost", "Profit"]
                title = f"Profit Analysis Report ({date_from} to {date_to})"
                
            elif report_type == 'supplier_performance':
                # Supplier Performance
                cur.execute("""
                    SELECT s.name, s.rating, s.lead_time_days,
                           COUNT(p.id) as product_count,
                           SUM(p.stock) as total_stock,
                           AVG(p.selling_price - p.purchase_price) as avg_margin
                    FROM suppliers s
                    LEFT JOIN products p ON s.id = p.supplier_id
                    WHERE s.is_active = 1
                    GROUP BY s.id, s.name, s.rating, s.lead_time_days
                    ORDER BY s.rating DESC
                """)
                columns = ["Supplier", "Rating", "Lead Time (days)", "Products", "Total Stock", "Avg Margin"]
                title = "Supplier Performance Report"
                
            elif report_type == 'inventory_valuation':
                # Inventory Valuation
                cur.execute("""
                    SELECT category,
                           COUNT(*) as item_count,
                           SUM(stock) as total_units,
                           SUM(stock * purchase_price) as total_cost,
                           SUM(stock * selling_price) as total_retail,
                           SUM(stock * (selling_price - purchase_price)) as potential_profit
                    FROM products
                    WHERE status = 'active'
                    GROUP BY category
                    ORDER BY total_cost DESC
                """)
                columns = ["Category", "Items", "Units", "Total Cost", "Total Retail", "Potential Profit"]
                title = "Inventory Valuation Report"
                
            elif report_type == 'slow_movers':
                # Slow Moving Items
                cur.execute("""
                    SELECT p.model, p.category, p.stock,
                           COUNT(s.id) as sales_count,
                           COALESCE(MAX(s.date), 'Never') as last_sale
                    FROM products p
                    LEFT JOIN sales s ON p.model = s.model
                    WHERE p.status = 'active'
                    GROUP BY p.model, p.category, p.stock
                    HAVING sales_count <= 2 OR last_sale < date('now', '-30 days')
                    ORDER BY last_sale ASC
                """)
                columns = ["Model", "Category", "Stock", "Sales Count", "Last Sale"]
                title = "Slow Moving Items Report"
                
            elif report_type == 'fast_movers':
                # Fast Moving Items
                cur.execute("""
                    SELECT p.model, p.category,
                           COUNT(s.id) as sales_count,
                           SUM(s.quantity) as total_sold,
                           SUM(s.total) as revenue
                    FROM products p
                    JOIN sales s ON p.model = s.model
                    WHERE s.date BETWEEN ? AND ?
                    GROUP BY p.model, p.category
                    ORDER BY total_sold DESC
                    LIMIT 20
                """, (date_from, date_to))
                columns = ["Model", "Category", "Sales Count", "Total Sold", "Revenue"]
                title = f"Top 20 Fast Moving Items ({date_from} to {date_to})"
                
            else:
                styled_label(parent, "Unknown report type", foreground=COLOR_DANGER).pack()
                return
            
            # Get data
            rows = cur.fetchall()
            
            # Display title
            styled_label(parent, title, font=FONT_BOLD).pack(anchor=tk.W, pady=(0, 10))
            
            if not rows:
                styled_label(parent, "No data found for selected criteria", foreground=COLOR_WARNING).pack()
                return
            
            # Create table
            table_frame = ttk.Frame(parent)
            table_frame.pack(fill="both", expand=True)
            
            # Treeview for data
            columns = tuple(col.lower().replace(' ', '_') for col in columns)
            tree = ttk.Treeview(table_frame, columns=columns, show="headings")
            
            for col, header in zip(columns, columns):
                tree.heading(col, text=col.replace('_', ' ').title())
                tree.column(col, width=100)
            
            # Scrollbar
            scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side=tk.RIGHT, fill="y")
            tree.pack(side=tk.LEFT, fill="both", expand=True)
            
            # Insert data
            for row in rows:
                values = []
                for val in row:
                    if isinstance(val, float):
                        values.append(f"{val:,.2f}")
                    elif val is None:
                        values.append("N/A")
                    else:
                        values.append(str(val))
                tree.insert("", "end", values=values)
            
            # Summary
            summary_frame = ttk.Frame(parent)
            summary_frame.pack(fill="x", pady=(10, 0))
            
            styled_label(summary_frame, f"Total Records: {len(rows)}", font=FONT_BOLD).pack(side=tk.LEFT)
            
            # Store data for export
            parent.report_data = {
                'title': title,
                'columns': columns,
                'rows': rows
            }
            
    except Exception as e:
        logging.exception("Failed to generate report")
        styled_label(parent, f"Error generating report: {e}", foreground=COLOR_DANGER).pack()


def export_report_to_excel(report_type: str, parent):
    """Export report to Excel."""
    if not EXCEL_AVAILABLE:
        messagebox.showerror("Error", "openpyxl library required for Excel export")
        return
    
    if not hasattr(parent, 'report_data'):
        messagebox.showinfo("Info", "Generate a report first")
        return
    
    data = parent.report_data
    
    # File dialog
    filename = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile=f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    
    if not filename:
        return
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = report_type
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = data['title']
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=16)
        
        # Headers
        headers = [col.replace('_', ' ').title() for col in data['columns']]
        ws.append(headers)
        
        # Data rows
        for row in data['rows']:
            ws.append([str(val) if val is not None else 'N/A' for val in row])
        
        # Style
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = min(adjusted_width, 50)
        
        wb.save(filename)
        messagebox.showinfo("Success", f"Report exported to:\n{filename}")
        
    except Exception as e:
        logging.exception("Excel export failed")
        messagebox.showerror("Error", f"Export failed: {e}")


def export_report_to_pdf(report_type: str, parent):
    """Export report to PDF."""
    if not PDF_AVAILABLE:
        messagebox.showerror("Error", "reportlab library required for PDF export")
        return
    
    if not hasattr(parent, 'report_data'):
        messagebox.showinfo("Info", "Generate a report first")
        return
    
    data = parent.report_data
    
    # File dialog
    filename = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")],
        initialfile=f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    )
    
    if not filename:
        return
    
    try:
        doc = SimpleDocTemplate(filename, pagesize=A4)
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=12
        )
        
        elements.append(Paragraph(data['title'], title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Generated date
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Table data
        table_data = [data['columns']]
        for row in data['rows'][:100]:  # Limit to 100 rows
            table_data.append([str(val) if val is not None else 'N/A' for val in row])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        messagebox.showinfo("Success", f"Report exported to:\n{filename}")
        
    except Exception as e:
        logging.exception("PDF export failed")
        messagebox.showerror("Error", f"Export failed: {e}")
